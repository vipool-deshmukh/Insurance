class Employee:

    def __init__(self,eid,enm,eag,esal,erole,eadr):
        self.empId = int(eid)
        self.empName = enm
        self.empAge = int(eag)
        self.empSalary = float(esal)
        self.empRole = erole
        self.empAddress = eadr

    def __str__(self):
        return f'''\n\n {self.__dict__}'''

    def __repr__(self):
        return str(self)

ADDRESSLIST = ['PUNE','MUMBAI','SATARA','SANGLI','KOLHAPUR','LATUR','NANDED','ANAGAR','BEED']
ROLESLIST = ['SSE','SE','MANAGER','CEO','STAFF','HOUSKEEPING','LEAD']

import random
emp_count = 770
def get_dummy_employee(n):
    global emp_count

    emplist = []
    for item in range(n):
        emp_count = emp_count + 1
        name = f'{chr((random.randint(65,90)))}AAAA'
        age = random.randint(22,60)
        salary = round(random.randint(30000,80000)/3,2)
        adr = ADDRESSLIST[random.randint(0,len(ADDRESSLIST)-1)]
        role = ROLESLIST[random.randint(0, len(ROLESLIST) - 1)]
        emplist.append(Employee(eid=emp_count,enm=name,eag=age,esal=salary,erole=role,eadr=adr))
    return emplist


emplist = get_dummy_employee(1000)
#print(emp)

                    #python - json,csv,sqlite3 db -->
import openpyxl     # openxl -- > 3rd party --> bcoz nt provided by python

import os
FILE_PATH = os.getcwd()+"\\empdata.xlsx"
from openpyxl.styles import Font
def write_data_into_excel():
    workbook = openpyxl.Workbook()  # CREATES NEW WORKBOOK
    sheet = workbook.create_sheet('empinfo')
    sheet['A1'] = 'EMP_ID'
    sheet['B1'] = 'EMP_NAME'
    sheet['C1'] = 'EMP_AGE'
    sheet['D1'] = 'EMP_SALARY'
    sheet['E1'] = 'EMP_ROLE'
    sheet['F1'] = 'EMP_ADDRESS'
    sheet['A1'].font = Font(bold=True)
    for index,emp in enumerate(emplist):
        rownum = index + 2
        sheet['A'+str(rownum)] = emp.empId
        sheet['B'+str(rownum)] = emp.empName
        sheet['C'+str(rownum)] = emp.empAge
        sheet['D'+str(rownum)] = emp.empSalary
        sheet['E'+str(rownum)] = emp.empRole
        sheet['F'+str(rownum)] = emp.empAddress

    workbook.save(FILE_PATH)


write_data_into_excel()


def read_data_from_excel():
    workbook = openpyxl.load_workbook(FILE_PATH)
    sheet = workbook['empinfo']
    nrows = sheet.max_row
    emplist = []
    for row in range(1,nrows+1):
        if row == 1:    # 1 -- header -> 1 row chi skip kartoy
            continue
        eid = sheet['A' + str(row)].value
        enm = sheet['B' + str(row)].value
        eag = sheet['C' + str(row)].value
        esal = sheet['D' + str(row)].value
        erole = sheet['E' + str(row)].value
        eadr = sheet['F' + str(row)].value
        emplist.append(Employee(eid,enm,eag,esal,erole,eadr))

    return emplist



emplist = read_data_from_excel()
print(emplist)

#yaml --> xml --> **