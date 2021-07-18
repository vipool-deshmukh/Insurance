import openpyxl

import random
class Product:

    def __init__(self,pid,pnm,prc,pqty,pven,remarks='NA'):
        self.prodId = int(pid)
        self.prodName = pnm
        self.prodQty = int(pqty)
        self.prodVendor = pven
        self.prodPrice = float(prc)
        self.prodRemarks = remarks

    def __repr__(self): # already object -->
        return str(self)

    def __str__(self):
        return f'''\n {self.__dict__}'''

    @classmethod
    def get_product_instance(cls,num):      # instances -- factory method --> logic- object creation- with the help cls name
        prodlist =[]
        for item in range(num):
            pid = int(input('Enter Product Id : '))
            pnm = input('Enter Product Name : ')
            pqty = int(input('Enter Product Qty : '))
            prc = float(input('Enter Product Price :'))
            pven = input('Enter Product Vendor : ')
            prodlist.append(cls(pid,pnm,prc,pqty,pven))
        return prodlist

    @classmethod
    def get_random_products(cls,num):  #10
        prodlist = []
        vendorlist =['AAAA','BBBB','CCCC','DDDD','EEEE']
        for item in range(num):             #0-9
            pid = item + 1
            pnm = f'{chr(random.randint(65,90))}AAAA'
            pqty = random.randint(1,20)
            prc = random.randint(1000,9000)
            pven = vendorlist[random.randint(0,len(vendorlist)-1)]
            if item%3==0:
                remarks = f'XXX{item}'
                prodlist.append(cls(pid, pnm, prc, pqty, pven,remarks))
            else:
                prodlist.append(cls(pid, pnm, prc, pqty, pven))

        return prodlist


import os
EXCEL_FILE_PATH = os.getcwd() + "\\products.xlsx"



from openpyxl.styles.borders import Border, Side
#
from openpyxl.styles import PatternFill
#sheet['A1'].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

def write_data_into_excel(products):        #products --> list first --> product instance
    workbook = openpyxl.Workbook()      # Workbook -->
    sheet = workbook.create_sheet('pdata')
    if products:
        #----headers -->
        firstproduct = products[0]
        columns = []
        for item in firstproduct.__dict__:          #,keys      items()     values
            columns.append(item.upper())  #attributes - means columns --name
        print(columns)
        firstcoln = 65
        for col in columns:         #column names -- means attribute        -->
            sheet[chr(firstcoln) + str(1)] = col
            sheet[chr(firstcoln) + str(1)].fill = PatternFill(start_color="FFC7CE", fill_type = "solid")
            sheet[chr(firstcoln) + str(1)].border = thin_border
            firstcoln = firstcoln + 1

        #--write data -->
        rownum = 2
        for prod in products:
            row = str(rownum)
            sheet['A'+row] = prod.prodId
            sheet['B'+row] = prod.prodName
            sheet['C'+row] = prod.prodQty
            sheet['D'+row] = prod.prodVendor
            sheet['E'+row] = prod.prodPrice
            sheet['F'+row] = prod.prodRemarks
            rownum = rownum + 1


    workbook.save(EXCEL_FILE_PATH)


prods = Product.get_random_products(1000)
write_data_into_excel(prods)



def read_data_from_excel():
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook['pdata']
    maxrow = sheet.max_row
    prodlist = []
    for row in range(1,maxrow+1):
        if row == 1:        # header row
            continue

        pid = sheet['A' + str(row)].value
        pnm = sheet['B' + str(row)].value
        pqty = sheet['C' + str(row)].value
        pven = sheet['D' + str(row)].value
        prc = sheet['E' + str(row)].value
        prm = sheet['F' + str(row)].value
        prodlist.append(Product(pid,pnm,prc,pqty,pven,prm))

    return prodlist


def delete_product(pid):
    prodlist = read_data_from_excel()
    for prod in prodlist:
        if prod.prodId == pid:
            prodlist.remove(prod)
            write_data_into_excel(prodlist)
            print('Removed...!')
            break


delete_product(777)

pdata = read_data_from_excel()
print(pdata)
